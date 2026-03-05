"""
File input/output adapters for lead integration and export.
"""

from __future__ import annotations

import csv
import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from scripts.leadgen.exceptions import EmptyExportError, InvalidSABIFormatError
from scripts.leadgen.models import CSV_HEADERS, Lead, Vertical
from scripts.leadgen.normalize import (
    classify_vertical,
    clean_value,
    extract_email,
    extract_linkedin,
)
from scripts.leadgen.sabi import (
    parse_sabi_revenue,
    resolve_vertical,
    strip_legal_suffix,
)

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Header synonym mapping — SABI exports vary across ESADE access versions
# ---------------------------------------------------------------------------
_SABI_HEADER_SYNONYMS: dict[str, list[str]] = {
    "legal_name": ["nombre", "razón social", "razon social", "denominación social"],
    "trade_name": ["nombre comercial", "trade name"],
    "nif": ["nif", "cif", "tax id"],
    "address": ["dirección domicilio social", "direccion domicilio social", "dirección"],
    "postal_code": ["código postal", "codigo postal", "cp"],
    "city": ["localidad", "ciudad", "city"],
    "province": ["provincia", "province"],
    "phone": ["teléfono", "telefono", "phone"],
    "website": ["página web", "pagina web", "web", "website", "url"],
    "cnae_primary": [
        "código cnae 2009 (primario)",
        "codigo cnae 2009 (primario)",
        "cnae primario",
        "cnae 2009",
        "cnae primary",
        "código cnae primario",
        "codigo cnae primario",
    ],
    "cnae_secondary": [
        "código cnae 2009 (secundario)",
        "codigo cnae 2009 (secundario)",
        "cnae secundario",
        "cnae secondary",
        "código cnae secundario",
        "codigo cnae secundario",
    ],
    "activity_description": [
        "descripción actividad",
        "descripcion actividad",
        "actividad",
        "activity description",
    ],
    "revenue_last": [
        "ingresos de explotación (eur mil) — último año",
        "ingresos de explotación (eur mil)",
        "ingresos de explotacion (eur mil)",
        "ingresos explotación mil eur",
        "ingresos explotacion mil eur",
        "cifra de negocios",
        "operating revenue",
        "ingresos de explotación",
        "ingresos de explotacion",
    ],
    "revenue_prev": [
        "ingresos de explotación (eur mil) — penúltimo año",
        "ingresos de explotacion (eur mil) — penultimo año",
        "revenue prev",
    ],
    "ebitda": [
        "ebitda (eur mil) — último año",
        "ebitda (eur mil)",
        "ebitda",
    ],
    "net_profit": [
        "resultado del ejercicio (eur mil) — último año",
        "resultado del ejercicio",
        "net profit",
    ],
    "employees": [
        "número empleados — último año",
        "numero empleados",
        "número empleados",
        "employees",
        "num empleados",
    ],
    "incorporation_date": [
        "fecha constitución",
        "fecha constitucion",
        "incorporation date",
    ],
    "company_status": ["estado", "status", "company status"],
    "legal_form": ["forma jurídica", "forma juridica", "legal form"],
}

# Statuses considered "active" (from architecture spec §3.4)
_ACTIVE_STATUSES: set[str] = {"activa", "activa - con incidencias"}

# Holding company CNAE
_HOLDING_CNAE = "6420"


def from_apollo_csv(path: Path) -> list[Lead]:
    """Parse raw Apollo.io CSV exports into a list of Lead objects.

    Ported directly from the original ingestion logic. Performed line-by-line
    reading with manual splitting to handle Apollo's specific formatting.

    Args:
        path (Path): The pathlib.Path to the Apollo CSV.

    Returns:
        list[Lead]: A list of assembled Lead objects (unverified).
    """
    leads: list[Lead] = []

    with open(path, encoding="utf-8-sig") as f:
        # Skip header
        next(f, None)

        for line_num, line in enumerate(f, start=2):
            if len(line.strip()) < 10:
                log.warning(
                    "Skipping line %d (line %d in file): Too short (<10 chars)",
                    len(leads) + 1,
                    line_num,
                )
                continue

            # Raw split since Apollo puts email at index 5 or 6 and quote structure is messy
            parts = line.split(",")

            first_name = clean_value(parts[0]) if len(parts) > 0 else ""
            last_name = clean_value(parts[1]) if len(parts) > 1 else ""
            name = f"{first_name} {last_name}".strip()

            company = clean_value(parts[3]) if len(parts) > 3 else ""

            email = extract_email(parts)
            if not email:
                log.warning(
                    "Skipping line %d (line %d in file): Unable to parse valid email address",
                    len(leads) + 1,
                    line_num,
                )
                continue

            linkedin = extract_linkedin(parts) or ""
            vertical = classify_vertical(line)

            lead = Lead(
                company_name=company,
                contact_name=name,
                email=email,
                revenue_est="€5M-€20M",  # Matches original hardcoded process_leads logic
                vertical=vertical,
                linkedin_url=linkedin,
                confidence_score=0,
                next_action="Research & Send Email #1",
            )
            leads.append(lead)
            log.info("Parsed %s (%s)", lead.company_name, lead.email)

    return leads


# ---------------------------------------------------------------------------
# SABI XLSX adapter
# ---------------------------------------------------------------------------


def _build_column_map(header_cells: tuple[Any, ...]) -> dict[str, int]:
    """Build a mapping from internal field names to column indices.

    Uses ``_SABI_HEADER_SYNONYMS`` to flexibly match header labels.
    """
    col_map: dict[str, int] = {}
    for idx, cell in enumerate(header_cells):
        if cell is None:
            continue
        normalised = str(cell).strip().lower()
        for field_name, synonyms in _SABI_HEADER_SYNONYMS.items():
            if normalised in synonyms and field_name not in col_map:
                col_map[field_name] = idx
                break
    return col_map


def _cell(row: tuple[Any, ...], col_map: dict[str, int], field: str) -> Any:
    """Safely extract a cell value by field name."""
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _format_revenue_est(revenue_eur: int | None) -> str:
    """Format an integer EUR revenue into a human-readable string."""
    if revenue_eur is None or revenue_eur == 0:
        return ""
    millions = revenue_eur / 1_000_000
    return f"€{millions:.1f}M"


def from_sabi_xlsx(
    path: Path,
    *,
    min_revenue_eur: int = 5_000_000,
    max_revenue_eur: int = 20_000_000,
    min_employees: int = 3,
    include_unverified_revenue: bool = False,
    cnae_map: dict[str, tuple[str, bool, str]] | None = None,
) -> list[Lead]:
    """Parse a SABI XLSX export into Lead objects.

    Reads a Bureau van Dijk SABI export file (XLSX format) and produces a
    list of company-level Lead stubs.  Each Lead contains verified financial
    data (revenue, EBITDA, employees) and vertical classification derived
    from CNAE codes, but does **not** contain individual contact names or
    email addresses — these must be enriched downstream.

    Args:
        path: Path to the SABI XLSX file.
        min_revenue_eur: Minimum revenue in EUR.  Companies below are excluded.
        max_revenue_eur: Maximum revenue in EUR.  Companies above are excluded.
        min_employees: Minimum employee count.
        include_unverified_revenue: If ``True``, include companies where
            SABI revenue is null.
        cnae_map: Custom CNAE-to-vertical mapping.  Falls back to
            :data:`CNAE_VERTICAL_MAP` when ``None``.

    Returns:
        Company-level Lead objects sorted by ``revenue_eur`` descending.

    Raises:
        FileNotFoundError: If *path* does not exist.
        InvalidSABIFormatError: If fewer than 5 expected columns are found.
        EmptyExportError: If zero leads remain after filtering.
    """
    if not path.exists():
        raise FileNotFoundError(path)

    # --- 1. Open workbook & map headers ---
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        wb.close()
        raise InvalidSABIFormatError("Workbook has no active sheet")

    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        raise InvalidSABIFormatError("XLSX file is empty") from None

    col_map = _build_column_map(header_row)

    if len(col_map) < 5:
        wb.close()
        raise InvalidSABIFormatError(
            f"Only {len(col_map)} of the expected SABI columns were found. "
            f"Matched: {sorted(col_map.keys())}"
        )

    log.info("SABI header mapping: %s", col_map)

    # --- 2. Iterate and filter rows ---
    leads: list[Lead] = []
    total_rows = 0
    skipped_status = 0
    skipped_cnae = 0
    skipped_revenue = 0
    skipped_employees = 0
    skipped_holding = 0

    for row in rows_iter:
        total_rows += 1

        # Status filter
        raw_status = _cell(row, col_map, "company_status")
        status_str = str(raw_status).strip().lower() if raw_status else ""
        if status_str and status_str not in _ACTIVE_STATUSES:
            skipped_status += 1
            continue

        # CNAE extraction
        raw_cnae_p = _cell(row, col_map, "cnae_primary")
        cnae_primary = str(raw_cnae_p).strip() if raw_cnae_p else ""
        raw_cnae_s = _cell(row, col_map, "cnae_secondary")
        cnae_secondary = str(raw_cnae_s).strip() if raw_cnae_s else ""

        # Holding company filter
        if cnae_primary == _HOLDING_CNAE:
            skipped_holding += 1
            continue

        # Activity description (for conditional CNAE logic)
        raw_desc = _cell(row, col_map, "activity_description")
        activity_desc = str(raw_desc).strip() if raw_desc else ""

        # Vertical resolution — also acts as CNAE inclusion filter
        vertical = resolve_vertical(cnae_primary, cnae_secondary, activity_desc)
        if vertical == Vertical.UNKNOWN:
            # Not in our target CNAE set
            skipped_cnae += 1
            continue

        # Revenue
        raw_revenue = _cell(row, col_map, "revenue_last")
        revenue_eur = parse_sabi_revenue(raw_revenue)

        if revenue_eur is None:
            if not include_unverified_revenue:
                skipped_revenue += 1
                continue
            revenue_verified = False
        else:
            revenue_verified = True
            if revenue_eur < min_revenue_eur or revenue_eur > max_revenue_eur:
                skipped_revenue += 1
                continue

        # Employees
        raw_emp = _cell(row, col_map, "employees")
        emp: int = 0
        if raw_emp is not None:
            try:
                emp = int(raw_emp)
            except (ValueError, TypeError):
                emp = 0
        if emp < min_employees:
            skipped_employees += 1
            continue

        # Holding heuristic: employees < 3 AND revenue > 5M
        if emp < 3 and revenue_eur is not None and revenue_eur > 5_000_000:
            skipped_holding += 1
            continue

        # --- Extract remaining fields ---
        raw_legal = _cell(row, col_map, "legal_name")
        legal_name = str(raw_legal).strip() if raw_legal else ""

        raw_trade = _cell(row, col_map, "trade_name")
        trade_name = str(raw_trade).strip() if raw_trade else ""

        company_name = trade_name if trade_name else legal_name
        company_name = strip_legal_suffix(company_name)

        raw_nif = _cell(row, col_map, "nif")
        nif = str(raw_nif).strip() if raw_nif else ""

        raw_web = _cell(row, col_map, "website")
        website_str = ""
        if raw_web:
            website_str = str(raw_web).strip()
            # Strip protocol & trailing slashes
            for prefix in ("https://", "http://", "www."):
                if website_str.lower().startswith(prefix):
                    website_str = website_str[len(prefix) :]
            website_str = website_str.rstrip("/")
            if ";" in website_str:
                website_str = website_str.split(";")[0].strip()

        raw_ebitda = _cell(row, col_map, "ebitda")
        ebitda_eur = parse_sabi_revenue(raw_ebitda) or 0

        raw_city = _cell(row, col_map, "city")
        city = str(raw_city).strip() if raw_city else ""

        raw_prov = _cell(row, col_map, "province")
        province = str(raw_prov).strip() if raw_prov else ""

        lead = Lead(
            company_name=company_name,
            vertical=vertical,
            revenue_est=_format_revenue_est(revenue_eur),
            next_action="Research & Send Email #1",
            # SABI-specific fields
            nif=nif,
            legal_name=legal_name,
            revenue_eur=revenue_eur or 0,
            revenue_verified=revenue_verified,
            ebitda_eur=ebitda_eur,
            employees=emp,
            website=website_str,
            city=city,
            province=province,
            cnae_primary=cnae_primary,
            cnae_secondary=cnae_secondary,
            source="sabi",
        )
        leads.append(lead)

    wb.close()

    # --- 3. Deduplicate by NIF ---
    seen_nifs: set[str] = set()
    deduped: list[Lead] = []
    dup_count = 0
    for lead in leads:
        if lead.nif and lead.nif in seen_nifs:
            dup_count += 1
            continue
        if lead.nif:
            seen_nifs.add(lead.nif)
        deduped.append(lead)

    # --- 4. Sort by revenue descending ---
    deduped.sort(key=lambda ld: ld.revenue_eur, reverse=True)

    # --- 5. Log summary ---
    log.info(
        "SABI import summary: total_rows=%d | skipped_status=%d | "
        "skipped_cnae=%d | skipped_revenue=%d | skipped_employees=%d | "
        "skipped_holding=%d | duplicates=%d | final_leads=%d",
        total_rows,
        skipped_status,
        skipped_cnae,
        skipped_revenue,
        skipped_employees,
        skipped_holding,
        dup_count,
        len(deduped),
    )

    if not deduped:
        raise EmptyExportError(
            f"SABI file {path.name} yielded 0 leads after filtering ({total_rows} rows processed)."
        )

    return deduped


def write_leads_csv(leads: list[Lead], path: Path) -> None:
    """Write Lead objects to a properly formatted CRM CSV file.

    Uses csv.DictWriter with the explicit module CSV_HEADERS to ensure
    perfect column alignment regardless of dataclass shifts.

    Args:
        leads (list[Lead]): The list of processed Lead objects.
        path (Path): Where to save the output CSV.
    """
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()

        for lead in leads:
            # Convert lead.to_row() list back to dict using CSV_HEADERS
            row_dict = dict(zip(CSV_HEADERS, lead.to_row(), strict=True))
            writer.writerow(row_dict)

    log.info("Wrote %d leads to %s", len(leads), path)
